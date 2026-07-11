#!/usr/bin/env python3
"""
合并指定目录下的所有 .xlsx 文件到一个工作簿。

规则：
- 每个源文件的内容放到以「文件名（不含扩展名）」命名的 sheet 中；
- 若某个源文件本身含多个 sheet，则子 sheet 命名为「文件名 - 原sheet名」；
- 自动处理 Excel sheet 名限制：非法字符替换、超 31 字符截断、重名加序号；
- 跳过输出文件本身、Excel 临时文件（~$ 开头）、隐藏文件。

用法：
    python merge_excels.py <目录> [-o 输出文件] [-r] [--values-only]

示例：
    python merge_excels.py ./reports
    python merge_excels.py ./reports -o all.xlsx -r
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter  # noqa: F401 (保留以便扩展)

# Excel sheet 名非法字符与长度上限
_INVALID_CHARS = set(r'\/*?:[]')
_MAX_SHEET_LEN = 31


def sanitize_sheet_name(name: str) -> str:
    """把任意字符串清洗成合法的 sheet 名片段（非法字符换成下划线，并去除首尾空白）。"""
    cleaned = "".join("_" if ch in _INVALID_CHARS else ch for ch in name).strip()
    # sheet 名不能以单引号开头/结尾
    cleaned = cleaned.strip("'")
    return cleaned or "Sheet"


def unique_sheet_name(base: str, used: set[str]) -> str:
    """在已用名集合内生成唯一且不超长的 sheet 名。"""
    base = sanitize_sheet_name(base)[:_MAX_SHEET_LEN]
    if base not in used:
        used.add(base)
        return base
    # 重名：追加 _2 / _3 ...，同时保证总长 <= 31
    idx = 2
    while True:
        suffix = f"_{idx}"
        candidate = base[: _MAX_SHEET_LEN - len(suffix)] + suffix
        if candidate not in used:
            used.add(candidate)
            return candidate
        idx += 1


def copy_sheet(src_ws, dst_ws, values_only: bool) -> None:
    """把源 sheet 的内容复制到目标 sheet。默认复制值+基础样式，values_only 时只复制值。"""
    for row in src_ws.iter_rows():
        for cell in row:
            if cell.value is None and not cell.has_style:
                continue
            new_cell = dst_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if not values_only and cell.has_style:
                new_cell.font = cell.font.copy()
                new_cell.fill = cell.fill.copy()
                new_cell.border = cell.border.copy()
                new_cell.alignment = cell.alignment.copy()
                new_cell.number_format = cell.number_format
    # 复制列宽
    for key, dim in src_ws.column_dimensions.items():
        if dim.width:
            dst_ws.column_dimensions[key].width = dim.width
    # 复制行高
    for key, dim in src_ws.row_dimensions.items():
        if dim.height:
            dst_ws.row_dimensions[key].height = dim.height
    # 复制合并单元格
    for merged in list(src_ws.merged_cells.ranges):
        dst_ws.merge_cells(str(merged))


def find_xlsx_files(directory: Path, recursive: bool, output: Path) -> list[Path]:
    """收集目录下待合并的 .xlsx，按文件名排序，排除临时/隐藏/输出文件自身。"""
    pattern = "**/*.xlsx" if recursive else "*.xlsx"
    files: list[Path] = []
    for p in directory.glob(pattern):
        if not p.is_file():
            continue
        if p.name.startswith("~$") or p.name.startswith("."):
            continue  # Excel 临时锁文件 / 隐藏文件
        if p.resolve() == output.resolve():
            continue  # 不要把输出文件也合并进去
        files.append(p)
    return sorted(files, key=lambda x: x.name.lower())


def merge(directory: Path, output: Path, recursive: bool, values_only: bool) -> int:
    files = find_xlsx_files(directory, recursive, output)
    if not files:
        print(f"⚠️  目录中没有找到可合并的 .xlsx 文件：{directory}")
        return 1

    dst_wb = openpyxl.Workbook()
    dst_wb.remove(dst_wb.active)  # 移除默认空 sheet
    used_names: set[str] = set()

    ok, failed = 0, 0
    for f in files:
        try:
            src_wb = openpyxl.load_workbook(f, data_only=values_only)
        except Exception as e:  # 单文件损坏不影响整体
            print(f"❌ 跳过（读取失败）：{f.name} —— {e}")
            failed += 1
            continue

        stem = f.stem  # 文件名（不含扩展名）
        visible_sheets = src_wb.worksheets
        multi = len(visible_sheets) > 1

        for ws in visible_sheets:
            base = f"{stem} - {ws.title}" if multi else stem
            sheet_name = unique_sheet_name(base, used_names)
            dst_ws = dst_wb.create_sheet(title=sheet_name)
            copy_sheet(ws, dst_ws, values_only)
            print(f"✅ {f.name}"
                  + (f" [{ws.title}]" if multi else "")
                  + f"  →  sheet「{sheet_name}」")
        ok += 1

    if not dst_wb.sheetnames:
        print("⚠️  没有任何内容被合并，未生成输出文件。")
        return 1

    output.parent.mkdir(parents=True, exist_ok=True)
    dst_wb.save(output)
    print("-" * 56)
    print(f"🎉 合并完成：{ok} 个文件成功"
          + (f"，{failed} 个失败" if failed else "")
          + f" → {output}（共 {len(dst_wb.sheetnames)} 个 sheet）")
    return 0


def main() -> None:
    parser = argparse.ArgumentParser(
        description="合并目录下所有 .xlsx 到一个工作簿，每个源文件占一个 sheet。"
    )
    parser.add_argument("directory", help="要扫描的目录")
    parser.add_argument("-o", "--output", default="merged.xlsx",
                        help="输出文件路径（默认 merged.xlsx）")
    parser.add_argument("-r", "--recursive", action="store_true",
                        help="递归扫描子目录")
    parser.add_argument("--values-only", action="store_true",
                        help="只复制单元格的值（公式取计算结果，忽略样式，速度更快）")
    args = parser.parse_args()

    directory = Path(args.directory)
    if not directory.is_dir():
        print(f"❌ 目录不存在：{directory}")
        sys.exit(2)

    sys.exit(merge(directory, Path(args.output), args.recursive, args.values_only))


if __name__ == "__main__":
    main()
